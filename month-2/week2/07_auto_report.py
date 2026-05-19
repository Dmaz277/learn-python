# import all library 
import pandas as pd
import openpyxl 
import matplotlib.pyplot as plt

# read csv and load data.csv to excel
data = pd.read_csv('data.csv')
df = pd.DataFrame(data)
df.to_excel('data.xlsx',index=False)
wb = openpyxl.load_workbook('data.xlsx')
ws1 = wb.active
ws1.title = 'Auto_Report'
ws2 = wb.create_sheet('Chart')

# make some fungtion
# fungtion for write data to excel in sheet 1
def write_excel():
    # import Font and PatternFill From openpyxl.styles library
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill

    # edit style
    ws1['A1'].font = Font(name='Arial', size=12, bold=True, color='000000')
    ws1['B1'].font = Font(name='Arial', size=12, bold=True, color='000000')
    ws1['C1'].font = Font(name='Arial', size=12, bold=True, color='000000')
    ws1['D1'].font = Font(name='Arial', size=12, bold=True, color='000000')

    # add color header
    ws1['A1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
    ws1['B1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
    ws1['C1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
    ws1['D1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
   
    # column
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['D'].width = 30

# function for make bar chart
def chart():
    # axes for x and y
    x = df['name_product']
    y = df['price']

    # make bar chart
    plt.bar(x,y)
    plt.savefig('chart.png')
    


# function for export chart to sheet 2
def export():
    from openpyxl.drawing.image import Image
    img = Image('chart.png')
    ws2.add_image(img,'A1')

# save the excel
def save():
    wb.save('auto_report.xlsx')

# Call All Function
write_excel()
chart()
export()
save()