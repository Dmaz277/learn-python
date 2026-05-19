# import all library that i had learned
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt

# read csv and load data from excel
data = pd.read_excel('data.xlsx')
data.to_excel('mini_project_report.xlsx',index=False)
wb = openpyxl.load_workbook('data.xlsx')
ws = wb.active
ws2 = wb.create_sheet('Chart')

# make some function
# fungtion for write data to excel in sheet 1
def write_excel():
    # import Font and PatternFill From openpyxl.styles library
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill

    # edit style
    ws['A1'].font = Font(name='Arial', size=12, bold=True, color='000000')
    ws['B1'].font = Font(name='Arial', size=12, bold=True, color='000000')
    ws['C1'].font = Font(name='Arial', size=12, bold=True, color='000000')
    ws['D1'].font = Font(name='Arial', size=12, bold=True, color='000000')

    # add color header
    ws['A1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
    ws['B1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
    ws['C1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
    ws['D1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')

    # add column dimesion as i want
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30

# function for conditional format,if something happen it will happening
def conditional_format():
    # import openpyxl to add color in looping
    from openpyxl.styles import PatternFill

    # looping for something condition
    for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
        if int(row[2].value) < 20 :
            for cell in row:
                cell.fill = PatternFill(start_color='FF0000',end_color='FF0000',fill_type='solid')

# function for make bar chart
def bar_chart():
    # make variable to save value chart
    x = data['name_product']
    y = data['price']

    # make the bar chart
    plt.bar(x,y)
    plt.savefig('chart2.png')

# function for export chart to sheet 2
def export():
    # impoer image
    from openpyxl.drawing.image import Image
    img = Image('chart2.png')
    ws2.add_image(img,'A1')

# save the excel
def save():
    wb.save('mini_project_report.xlsx')

write_excel()
conditional_format()
bar_chart()
export()
save()