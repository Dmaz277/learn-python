# library for edit display excel
import openpyxl 
# workbook = for open completely excel file
wb = openpyxl.load_workbook('data.xlsx')
# worksheet = one sheet in excel active
ws = wb.active

# edit styles
from openpyxl.styles import Font
ws['A1'].font = Font(bold=True)
ws['B1'].font = Font(bold=True)
ws['C1'].font = Font(bold=True)
ws['D1'].font = Font(bold=True)


# add color
from openpyxl.styles import PatternFill
ws['A1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
ws['B1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
ws['C1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
ws['D1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')

# column witdh
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20

# complete version
#from openpyxl.styles import Font
#ws['A1'].font = Font(name='Arial', size=12, bold=True, color='FFFFFF')

wb.save('data.xlsx')