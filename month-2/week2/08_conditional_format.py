# import library 
import pandas as pd
import openpyxl

# read csv and load data.csv to excel
data = pd.read_csv('data.csv')
df = pd.DataFrame(data)
df.to_excel('data.xlsx',index=False)
wb = openpyxl.load_workbook('data.xlsx')
ws = wb.active
ws.title = 'conditional_format'

# make function write excel and save
def write_excel():
    # import Font and PatternFill From openpyxl.styles library
    from openpyxl.styles import Font
    from openpyxl.styles import PatternFill

    # edit style
    ws['A1'].font = Font(name='Arial', size=12, bold=True, color='000000')
    ws['B1'].font = Font(name='Arial', size=12, bold=True, color='000000')
    ws['C1'].font = Font(name='Arial', size=12, bold=True, color='000000')
    ws['D1'].font = Font(name='Arial', size=12, bold=True, color='000000')

    # add color header
    ws['A1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
    ws['B1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
    ws['C1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
    ws['D1'].fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')
   

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30

    #loop for if a condition occurs
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if int(row[2].value) < 20 :
            for cell in row :
                cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')


def save():
    wb.save('conditional_format.xlsx')

# call all function
write_excel()
save()
